import csv
import io
from datetime import date
from sqlalchemy.orm import Session
from sqlalchemy import select, func, desc

from app.models.orden import Orden
from app.models.ejecucion import Ejecucion
from app.models.posicion import Posicion
from app.models.comision import Comision
from app.models.precio_mercado import PrecioMercado


def _apply_header_style(ws, headers: list[str]) -> None:
    """Applies dark header styling and freezes the first row."""
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.append(headers)
    fill = PatternFill(start_color="1E2A38", end_color="1E2A38", fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=10)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"


def _autofit_columns(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)


def get_summary(db: Session) -> dict:
    hoy = date.today()

    fills_hoy = db.execute(
        select(func.count(Ejecucion.id)).where(Ejecucion.fecha == hoy)
    ).scalar() or 0

    ordenes_total = db.execute(
        select(func.count(Orden.id))
    ).scalar() or 0

    ordenes_ejecutadas = db.execute(
        select(func.count(Orden.id)).where(Orden.estado_color == "green")
    ).scalar() or 0

    ordenes_pendientes = db.execute(
        select(func.count(Orden.id)).where(Orden.estado_color == "orange")
    ).scalar() or 0

    ordenes_error = db.execute(
        select(func.count(Orden.id)).where(Orden.estado_color == "red")
    ).scalar() or 0

    vol_result = db.execute(
        select(
            Orden.moneda,
            func.sum(Ejecucion.cantidad * Ejecucion.precio),
        )
        .join(Ejecucion, Ejecucion.orden_id == Orden.id)
        .group_by(Orden.moneda)
    ).all()
    volumen_por_moneda = {row[0]: round(row[1] or 0, 2) for row in vol_result}

    top_result = db.execute(
        select(
            Orden.especie,
            func.sum(Ejecucion.cantidad * Ejecucion.precio).label("volumen"),
            func.count(Ejecucion.id).label("fills"),
        )
        .join(Ejecucion, Ejecucion.orden_id == Orden.id)
        .group_by(Orden.especie)
        .order_by(desc("volumen"))
        .limit(5)
    ).all()

    return {
        "fills_hoy": fills_hoy,
        "ordenes_total": ordenes_total,
        "ordenes_ejecutadas": ordenes_ejecutadas,
        "ordenes_pendientes": ordenes_pendientes,
        "ordenes_error": ordenes_error,
        "volumen_por_moneda": volumen_por_moneda,
        "top_especies": [
            {"especie": r[0], "volumen": round(r[1] or 0, 2), "fills": r[2]}
            for r in top_result
        ],
    }


def get_concentracion(db: Session) -> dict:
    """
    Returns portfolio concentration by especie.
    valor_mercado = cantidad_neta * precio_mercado (fallback: costo_promedio_compra).
    """
    posiciones = db.execute(select(Posicion)).scalars().all()
    precios_map = {
        pm.especie: pm.precio
        for pm in db.execute(select(PrecioMercado)).scalars().all()
        if pm.precio
    }

    items = []
    for p in posiciones:
        neta = p.cantidad_neta or 0
        if neta <= 0:
            continue
        precio = precios_map.get(p.especie) or (p.costo_promedio_compra or 0)
        valor = round(neta * precio, 2)
        if valor > 0:
            items.append({
                "especie": p.especie,
                "moneda": p.moneda,
                "cantidad_neta": neta,
                "precio": precio,
                "valor_mercado": valor,
                "tiene_precio_mercado": p.especie in precios_map,
                "pct": 0.0,
            })

    # Aggregate by especie+moneda (sum across clients)
    agg: dict[tuple, dict] = {}
    for it in items:
        key = (it["especie"], it["moneda"])
        if key not in agg:
            agg[key] = {**it}
        else:
            agg[key]["valor_mercado"] = round(agg[key]["valor_mercado"] + it["valor_mercado"], 2)
            agg[key]["cantidad_neta"] += it["cantidad_neta"]

    result = sorted(agg.values(), key=lambda x: x["valor_mercado"], reverse=True)
    total = sum(r["valor_mercado"] for r in result)
    for r in result:
        r["pct"] = round(r["valor_mercado"] / total * 100, 2) if total > 0 else 0.0

    return {"items": result, "total_valor_mercado": round(total, 2)}


def get_benchmark(db: Session) -> dict:
    """
    Portfolio return vs S&P Merval since portfolio inception.

    Portfolio return: (valor_mercado_actual - costo_base) / costo_base * 100
    Merval return:    fetched via yfinance ^MERV from inception date to today
    Alpha:            portfolio_return - merval_return
    """
    # Inception date: earliest order
    inception = db.execute(select(func.min(Orden.fecha_orden))).scalar()
    if not inception:
        return {"error": "Sin órdenes en el sistema"}

    # Portfolio metrics
    posiciones = db.execute(select(Posicion)).scalars().all()
    precios_map = {
        pm.especie: pm.precio
        for pm in db.execute(select(PrecioMercado)).scalars().all()
        if pm.precio
    }

    costo_base = 0.0
    valor_actual = 0.0
    for p in posiciones:
        neta = p.cantidad_neta or 0
        comprada = p.cantidad_comprada or 0
        costo_pc = p.costo_promedio_compra or 0.0
        precio_mercado = precios_map.get(p.especie) or costo_pc

        costo_base += comprada * costo_pc
        valor_actual += max(neta, 0) * precio_mercado

    portfolio_pnl = round(valor_actual - costo_base, 2)
    portfolio_return = round(portfolio_pnl / costo_base * 100, 2) if costo_base > 0 else None

    # Merval return via yfinance (blocking — call inside sync route or to_thread)
    merval_return = None
    merval_inicio = None
    merval_actual = None
    try:
        import yfinance as yf
        hist = yf.Ticker("^MERV").history(start=inception.isoformat(), period="max")
        if not hist.empty:
            merval_inicio = float(hist["Close"].iloc[0])
            merval_actual = float(hist["Close"].iloc[-1])
            merval_return = round((merval_actual - merval_inicio) / merval_inicio * 100, 2)
    except Exception as e:
        print(f"[Benchmark] yfinance ^MERV error: {e}")

    alpha = None
    if portfolio_return is not None and merval_return is not None:
        alpha = round(portfolio_return - merval_return, 2)

    return {
        "inception_date":    inception.isoformat(),
        "costo_base":        round(costo_base, 2),
        "valor_actual":      round(valor_actual, 2),
        "portfolio_pnl":     portfolio_pnl,
        "portfolio_return":  portfolio_return,
        "merval_inicio":     round(merval_inicio, 2) if merval_inicio else None,
        "merval_actual":     round(merval_actual, 2) if merval_actual else None,
        "merval_return":     merval_return,
        "alpha":             alpha,
    }


def get_positions_snapshot(db: Session) -> list[dict]:
    """Aggregate positions by especie across all clients."""
    result = db.execute(
        select(
            Posicion.especie,
            Posicion.moneda,
            func.sum(Posicion.cantidad_neta).label("cantidad_neta_total"),
            func.count(Posicion.id).label("clientes_count"),
            func.sum(Posicion.cantidad_comprada * Posicion.costo_promedio_compra).label("valor_compra"),
            func.sum(Posicion.cantidad_comprada).label("total_comprado"),
        )
        .group_by(Posicion.especie, Posicion.moneda)
        .order_by(Posicion.especie)
    ).all()

    snapshot = []
    for r in result:
        costo_ponderado = (
            round(r.valor_compra / r.total_comprado, 4)
            if r.total_comprado and r.total_comprado > 0
            else 0.0
        )
        snapshot.append({
            "especie": r.especie,
            "moneda": r.moneda,
            "cantidad_neta_total": r.cantidad_neta_total or 0,
            "clientes_count": r.clientes_count,
            "costo_promedio_ponderado": costo_ponderado,
        })
    return snapshot


def export_csv(db: Session, tipo: str, fecha: date | None = None) -> str:
    output = io.StringIO()

    if tipo == "ejecuciones":
        writer = csv.writer(output)
        writer.writerow([
            "id", "fecha", "nro_orden", "especie", "cliente",
            "tipo_orden", "mercado", "cantidad", "precio", "importe", "nro_secuencia",
        ])
        stmt = (
            select(Ejecucion, Orden)
            .join(Orden, Ejecucion.orden_id == Orden.id)
            .order_by(Ejecucion.fecha.desc(), Ejecucion.id.desc())
        )
        if fecha:
            stmt = stmt.where(Ejecucion.fecha == fecha)
        for ejec, orden in db.execute(stmt).all():
            writer.writerow([
                ejec.id, ejec.fecha.strftime("%d/%m/%Y"),
                orden.nro_orden, orden.especie, orden.cliente, orden.tipo_orden,
                ejec.mercado, ejec.cantidad, ejec.precio,
                round(ejec.cantidad * ejec.precio, 2), ejec.nro_secuencia,
            ])

    elif tipo == "posiciones":
        writer = csv.writer(output)
        writer.writerow([
            "id", "cliente", "especie", "moneda", "mercado",
            "cantidad_comprada", "cantidad_vendida", "cantidad_neta",
            "costo_promedio_compra", "costo_promedio_venta",
        ])
        for p in db.execute(
            select(Posicion).order_by(Posicion.cliente, Posicion.especie)
        ).scalars().all():
            writer.writerow([
                p.id, p.cliente, p.especie, p.moneda, p.mercado,
                p.cantidad_comprada, p.cantidad_vendida, p.cantidad_neta,
                p.costo_promedio_compra, p.costo_promedio_venta,
            ])

    return output.getvalue()


def export_xlsx(db: Session) -> bytes:
    """
    Exports Ejecuciones, Posiciones, and Órdenes as a multi-sheet XLSX workbook.
    Requires openpyxl (pip install openpyxl).
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError(
            "openpyxl no está instalado. Ejecutar: pip install openpyxl"
        )

    wb = openpyxl.Workbook()

    # ── Sheet 1: Ejecuciones ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Ejecuciones"
    _apply_header_style(ws1, [
        "ID", "Fecha", "Nro Orden", "Especie", "Cliente", "Tipo",
        "Mercado", "Cantidad", "Precio", "Importe",
        "Comisión + IVA", "Costo Efectivo/Unit.",
    ])
    for ejec, orden in db.execute(
        select(Ejecucion, Orden)
        .join(Orden, Ejecucion.orden_id == Orden.id)
        .order_by(Ejecucion.fecha.desc(), Ejecucion.id.desc())
    ).all():
        com = db.execute(
            select(Comision).where(Comision.ejecucion_id == ejec.id)
        ).scalar_one_or_none()
        ws1.append([
            ejec.id,
            ejec.fecha.strftime("%d/%m/%Y") if ejec.fecha else "",
            orden.nro_orden, orden.especie, orden.cliente, orden.tipo_orden,
            ejec.mercado, ejec.cantidad, ejec.precio,
            round(ejec.cantidad * ejec.precio, 2),
            round(com.monto_total, 2) if com else None,
            round(com.costo_efectivo_unitario, 4) if com else None,
        ])
    _autofit_columns(ws1)

    # ── Sheet 2: Posiciones ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Posiciones")
    _apply_header_style(ws2, [
        "ID", "Cliente", "Especie", "Moneda", "Mercado",
        "Comprado", "Vendido", "Neto",
        "Costo Prom. Compra", "Costo Prom. Venta",
    ])
    for p in db.execute(
        select(Posicion).order_by(Posicion.cliente, Posicion.especie)
    ).scalars().all():
        ws2.append([
            p.id, p.cliente, p.especie, p.moneda, p.mercado,
            p.cantidad_comprada, p.cantidad_vendida, p.cantidad_neta,
            round(p.costo_promedio_compra, 4),
            round(p.costo_promedio_venta, 4),
        ])
    _autofit_columns(ws2)

    # ── Sheet 3: Órdenes ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Órdenes")
    _apply_header_style(ws3, [
        "Nro Orden", "Tipo", "Fecha", "Cliente", "Razón Social",
        "Especie", "Moneda", "Precio Límite", "Cant. Total",
        "Cant. Ejecutada", "Precio Prom.", "Progreso %", "Instancia",
    ])
    for o in db.execute(select(Orden).order_by(Orden.nro_orden)).scalars().all():
        progreso = round(o.cantidad_ejecutada / o.cantidad_total * 100, 1) if o.cantidad_total > 0 else 0
        ws3.append([
            o.nro_orden, o.tipo_orden,
            o.fecha_orden.strftime("%d/%m/%Y") if o.fecha_orden else "",
            o.cliente, o.razon_social, o.especie, o.moneda,
            o.precio_limite, o.cantidad_total, o.cantidad_ejecutada,
            o.precio_promedio, progreso, o.instancia,
        ])
    _autofit_columns(ws3)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
